"""
MongoDB-based temporary student reference database.
Students are uploaded via Excel before scanning and used to look up
names/emails by registration number during the scan flow.
"""
import io
from typing import List, Optional, Dict
from pymongo import MongoClient
from pymongo.errors import ConnectionFailure
from openpyxl import load_workbook
from app.config import settings


class StudentDBService:
    def __init__(self):
        self.client = None
        self.db = None
        self._connect()

    def _connect(self):
        if not settings.MONGODB_URL:
            print("WARNING: MONGODB_URL not set — student DB disabled")
            return
        try:
            self.client = MongoClient(settings.MONGODB_URL, serverSelectionTimeoutMS=5000)
            self.client.admin.command("ping")
            self.db = self.client[settings.MONGODB_DB_NAME]
            print(f"MongoDB connected: {settings.MONGODB_DB_NAME}")
        except ConnectionFailure as e:
            print(f"WARNING: MongoDB connection failed: {e}")
            self.client = None
            self.db = None

    def _collection(self, session_id: str):
        """Each scan session gets its own collection for isolation."""
        if self.db is None:
            return None
        return self.db[f"students_{session_id}"]

    def upload_excel(self, session_id: str, file_bytes: bytes, filename: str) -> Dict:
        """Parse an Excel/CSV file and store students in MongoDB for this session."""
        col = self._collection(session_id)
        if col is None:
            return {"success": False, "message": "MongoDB not connected", "count": 0}

        try:
            wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return {"success": False, "message": "Empty file", "count": 0}

            # First row = headers — normalize to lowercase
            raw_headers = [str(h).strip().lower() if h else "" for h in rows[0]]

            # Map common header variations
            header_map = {}
            for i, h in enumerate(raw_headers):
                if any(k in h for k in ["reg", "roll", "number", "id", "register"]):
                    header_map["register_number"] = i
                elif any(k in h for k in ["name", "student"]):
                    if "register_number" not in header_map or "name" not in raw_headers[header_map.get("register_number", -1)]:
                        header_map.setdefault("student_name", i)
                elif any(k in h for k in ["email", "mail"]):
                    header_map["email"] = i
                elif any(k in h for k in ["section", "sec"]):
                    header_map["section"] = i
                elif any(k in h for k in ["branch", "dept", "department"]):
                    header_map["branch"] = i
                elif any(k in h for k in ["year"]):
                    header_map["year"] = i

            if "register_number" not in header_map:
                return {"success": False, "message": "No registration number column found in headers", "count": 0}

            # Clear old data for this session
            col.delete_many({})

            docs = []
            for row in rows[1:]:
                reg = str(row[header_map["register_number"]] or "").strip()
                if not reg:
                    continue
                doc = {
                    "register_number": reg.upper(),
                    "student_name": str(row[header_map.get("student_name", 0)] or "").strip() if "student_name" in header_map else "",
                    "email": str(row[header_map.get("email", 0)] or "").strip() if "email" in header_map else "",
                    "section": str(row[header_map.get("section", 0)] or "").strip() if "section" in header_map else "",
                    "branch": str(row[header_map.get("branch", 0)] or "").strip() if "branch" in header_map else "",
                    "year": str(row[header_map.get("year", 0)] or "").strip() if "year" in header_map else "",
                }
                docs.append(doc)

            if docs:
                col.insert_many(docs)

            wb.close()
            return {"success": True, "message": f"{len(docs)} student(s) loaded", "count": len(docs)}
        except Exception as e:
            return {"success": False, "message": str(e), "count": 0}

    def find_student(self, session_id: str, register_number: str) -> Optional[Dict]:
        """Look up a student by reg number in this session's collection."""
        col = self._collection(session_id)
        if col is None:
            return None
        doc = col.find_one({"register_number": register_number.strip().upper()})
        if doc:
            doc.pop("_id", None)
            return doc
        return None

    def get_all_students(self, session_id: str) -> List[Dict]:
        """Get all students for a session."""
        col = self._collection(session_id)
        if col is None:
            return []
        docs = list(col.find({}, {"_id": 0}))
        return docs

    def get_student_count(self, session_id: str) -> int:
        col = self._collection(session_id)
        if col is None:
            return 0
        return col.count_documents({})

    def cleanup_session(self, session_id: str):
        """Drop the session's student collection when session ends."""
        if self.db is None:
            return
        try:
            self.db.drop_collection(f"students_{session_id}")
        except Exception:
            pass


student_db = StudentDBService()
